#######################################################################################################################
#######################################################################################################################
# Begin Main Code
#
# script to grade the SER401 / 402 individual contribution reports
# original script developed by Doug Sandy
# modification history:
# 200709 - VS - Added def date for end date as 13 days from start date for easy of use.
# 200709 v2.0 - VS - Scan all valid rows in the sheet instead of shutting down after the first blank line.
# 201013 v2.1 - VS - Added ability to have artifacts in google docs and fixed bug when evidence used twice - should count for one
# 201016 v2.2 - VS - Allowed in case % put in the %complete column.
#               VS - add print out message to sheet - msgs (for help in debug)


import openpyxl as xl
import urllib.request
import easygui
import os
import glob
import datetime
import tkinter as tk
from tkinter import ttk
from tkcalendar import Calendar
from Task import Task

def get_number_from_cell(cell):
    #################################################################################
    # getNumberFromCell
    #
    # Given a worksheet cell, returns a valid tuple where the first value is true
    # if the cell contains a number, or false otherwise.  The second value of the
    # tuple is the numeric value of the cell.
    #
    aFloat = 0.0
    try:
        aFloat = float(cell.value)
    except:
        return False, 0

    #if cell.value is None:
    #    return False, 0
    #if type(cell.value) is int or type(cell.value) is float:
    #    return True, cell.value
    #return False, 0

    return True, aFloat

def get_percent(w, r, c):
    ################################################################################
    # getPercent
    # return the percent value of the cell
    #
    t = str(w[r][c].value)
    w[r][c].value = t.replace("%", "")
    [result, num] = get_number_from_cell(w[r][c])
    if num <= 1:
        num = num * 100
    return num


def get_url(w, r, c):
    # return the string representation of the URL in the specified worksheet cell
    if w[r][c].hyperlink is None:
        return str(w[r][c].value)
    else:
        return str(w[r][c].hyperlink.target)


def get_submission_date(w, r, c, log):
    # returns a date structure for the date in the specified worksheet cell
    val = str(w[r][c].value)
    datevalid = 0
    try:
        result = datetime.datetime.strptime(val, "%Y-%m-%d %H:%M:%S").date()
        datevalid = 1
    except:
        datevalid = 0
    if datevalid == 0:
        try:
            result = datetime.datetime.strptime(val, "%m/%d/%Y").date()
            datevalid = 1
        except:
            datevalid = 0
    if datevalid == 0:
        if "One or more invalid dates." not in log:
            log = log + " One or more invalid dates."
        result = datetime.datetime.strptime("1900-01-01 00:00:00", "%Y-%m-%d %H:%M:%S").date()
    return log, result

def is_url_valid(url, log):
    ###############################################################################
    # isUrlValid
    #
    #  return True if the URL is valid, otherwise, false
    #
    # parameters:
    #    url - the url to check

    result = True
    url = url.lower()

    # url is not valid if it is a Github URL but does not include a commit
    if "github" in url and "commit" not in url:
        if "One or more Github links don't include commit information." not in log:
            log = log + " One or more Github links don't include commit information."
        return False, log
    if "gitlab" in url and "commit" not in url:
        if "One or more Gitlab links don't include commit information." not in log:
            log = log + " One or more Gitlab links don't include commit information."
        return False, log

    # typically allow early in the class - first couple of sprints.
    allowGoogleDocLink = True
    # allow for google docs to be valid
    if allowGoogleDocLink == True:
        if url.find('docs.google.com') or url.find('drive.gooogle.com'):
            result= True
    try:
        urllib.request.urlopen(url)
    except:
        result = False

    # the URL still may be valid if the site requires authentication
    if url.find('taiga.io') or url.find('jira'):
        result = True

    if not result:
        if "Unable to open one or more links." not in log:
            log = log + " Unable to open one or more links."
    return result, log


def is_ws_row_blank(w, r, max_column):
    ###############################################################################
    # isWsRowBlank
    #
    #  return True if the worksheet row is blank, otherwise return false.
    #
    # parameters:
    #    ws - a openpyxl worksheet object
    #    row - the row within the spreadsheet to check
    #    maxcolumn - the maximum column in the row to check
    for col in range(max_column):
        cell = w[r][col]
        if cell.hyperlink is None:
            if cell.value is not None:
                return False
        else:
            if cell.hyperlink.target is not None:
                return False
    return True


def get_date(message, defDate = datetime.datetime.now() ):

    # Prompt the user for a date using a small calendar widget.  The date value selected will be returned.

    # handler for calendar date selection
    def day_sel(*args):
        cal.root.destroy()
    top = tk.Tk()
    ttk.Label(top, text=message).pack(padx=10, pady=10)
    #now = datetime.datetime.now()
    #current_year = now.year
    #current_month = now.month
    #current_day = now.day
    defYear = defDate.year
    defMonth = defDate.month
    defDay = defDate.day

    cal = Calendar(top, font="Arial 14", selectmode='day', locale='en_US', cursor="hand1",
                   year=defYear, month=defMonth, day=defDay )
    cal.root = top
    cal.bind('<<CalendarSelected>>', day_sel)
    cal.pack(fill="both", expand=True)
    top.mainloop()
    return cal.selection_get()


########################################################################################################
########################################################################################################
# Start of main code

print()
print("CAPSTONE Individual Contribution Spreadsheet Autograder v2.2")

table_start_row = 16
table_end_row = 91

# get the input file folder
path = easygui.diropenbox()

# change the current working directory to the selected path
os.chdir(path)

# delete the output file if it exists
outfile = path+"//results.xlsx"
if os.path.exists(outfile):
    os.remove(outfile)

# get a list of all student files in the specified path
file_list = glob.glob(path+"//*.xlsx")
print("# of Submissions: " + str(len(file_list)))

# open the output workbook
owb = xl.Workbook()
owsData = owb.create_sheet('Data')
owsMsgs = owb.create_sheet('Msgs')

# initialize the result matrix
owsData['A1'] = 'Student'
owsData['B1'] = 'Overall Status'
owsData['C1'] = 'Technical Status'
owsData['D1'] = 'Teamwork Status'
owsData['E1'] = 'Sponsor Engagement'
owsData['F1'] = 'Score'
owsData['G1'] = 'Tasks Completed'
owsData['H1'] = 'Contribution Days'
owsData['I1'] = 'Largest Gap'
owsData['J1'] = 'Comments'

owsMsgs['A1'] = 'Submission'
owsMsgs['B1'] = 'Type'
owsMsgs['C1'] = 'Msg'

# get an assignment start and due dates
date_start = get_date('Assignment Start Date')
# let default date be 14 days from starting date
date_end = get_date('Assignment Due Date', date_start + datetime.timedelta(13))
period_len = date_end.toordinal() - date_start.toordinal() + 1
print("Start: ",  date_start)
print("End: ",  date_end)
print("Num days: ", period_len)

badurlindex = 0
badurl_list = {}
baddateindex = 0
baddate_list = {}
output_line = 2
msg_line = 2

# Loop for each individual contribution file
for filename in file_list:
    # clear the error log
    err_log = ""

    # get the student name - this is the filename up to the first '-' character
    name = filename[len(path)+1:]
    name = name[:name.find('_')]
    name = name.lower()

# moved from below - vbs
    print("%s\n", filename)
    # open the excel spreadsheet
    iwb = xl.load_workbook(filename)

    # select the main sheet of the workbook
    if 'Sheet1' in iwb.sheetnames:
        ws = iwb['Sheet1']
    else:
        print("Warning: " + filename + " does not have the default grading sheet of Sheet1 - Skipping submission.")
        continue

 #  vbs  print('%s\n', filename)
    tidx = 0
    t = {}
    row = table_start_row
# vbs now loop thru all rows and allow blank rows in between
    for row in range(table_start_row, table_end_row+1):
        if not is_ws_row_blank(ws, row, 4):
            # Process the input file task lines - create a list of contributions
            # in order to count as a contribution, it must have a
            # submission link, a submission date, and a non - zero contribution
            # percentage

            # ignore rows without contributions
            if len(get_url(ws, row, 0)) == 0:
                owsMsgs['A' + str(msg_line)] = filename
                owsMsgs['B' + str(msg_line)] = 'Missing Task/row skipped'
                owsMsgs['C' + str(msg_line)] = row
                msg_line += 1

                if "One or more tasks are not filled out." not in err_log:
                    err_log = err_log + " One or more tasks are not filled out."
                continue
            task = Task()
            task.urlMissing = False

            # get the task status
            task.valid = True
            status = str(ws[row][2].value)
            if 'complete' in status.lower():
                task.complete = True
            else:
                task.complete = False

            # check the url for the commit
            task.url = get_url(ws, row, 3)
            if len(task.url) < 5:
                owsMsgs['A' + str(msg_line)] = filename
                owsMsgs['B' + str(msg_line)] = 'Evidence link too short/missing'
                owsMsgs['C' + str(msg_line)] = task.url
                msg_line += 1

                if "Evidence link too short/missing." not in err_log:
                    err_log = err_log + " Evidence link too short/missing."
                task.urlValid = False
            else:
                [result, err_log]  = is_url_valid(task.url,err_log)
                if not result:
                    owsMsgs['A' + str(msg_line)] = filename
                    owsMsgs['B' + str(msg_line)] = 'Bad Evidence URL'
                    owsMsgs['C' + str(msg_line)] = task.url
                    msg_line += 1

                    badurl_list[badurlindex] = task.url
                    badurlindex = badurlindex + 1
                    task.urlValid = False
                else:
                    task.urlValid = True

            # get the date of the assignment relative to the grading period
            task.dateValid = True
            task.dateBadRange = False
            [err_log, assignment_date] = get_submission_date(ws, row, 4,err_log)
            if assignment_date > date_end:
                task.dateBadRange = True
                task.dateValid = False
            if assignment_date < date_start:
                task.dateBadRange = True
                task.dateValid = False
            delta = assignment_date - date_start
            task.dateidx = delta.days
            if task.dateBadRange is True:
                if " One or more dates are outside the assessment period" not in err_log:
                    err_log = err_log + " One or more dates are outside the assessment period"
                    err_log = err_log + " (" + date_start.strftime("%m/%d/%Y")+" - "+date_end.strftime("%m/%d/%Y")+")."

            task.percentWork = get_percent(ws, row, 5)

            # check the task link
            task.taskLink = get_url(ws, row, 0)
            if len(task.taskLink) < 5:
                owsMsgs['A' + str(msg_line)] = filename
                owsMsgs['B' + str(msg_line)] = ' Task link too short'
                owsMsgs['C' + str(msg_line)] = task.taskLink
                msg_line += 1

                if "One or more task links are invalid." not in err_log:
                    err_log = err_log + " One or more task links are invalid."
                task.taskLinkMissing = True
                task.taskLinkValid = False
            else:
                [result, err_log]  = is_url_valid(task.taskLink, err_log)
                if not result:
                    owsMsgs['A' + str(msg_line)] = filename
                    owsMsgs['B' + str(msg_line)] = 'Bad task URL'
                    owsMsgs['C' + str(msg_line)] = task.taskLink
                    msg_line += 1

                    badurl_list[badurlindex] = task.taskLink
                    badurlindex = badurlindex + 1
                    task.taskLinkValid = False
                    task.taskLinkMissing = False
                else:
                    task.taskLinkMissing = False
                    task.taskLinkValid = True

            t[tidx] = task
            tidx = tidx + 1
            row = row+1
        # END - go on to next loop

    # find any tasks that have duplicate commit urls - keep only the last date
    for i in range(tidx):
        for j in range(i):
            if t[i].urlValid and t[j].urlValid and t[i].url is t[j].url and t[i].dateidx != t[j].dateidx:
                # here for duplicate url
                if "Evidence link(s) used for more than one day(s)." not in err_log:
                    err_log = err_log + " Evidence link(s) used for more than one day(s)."
                    if t[i].dateidx > t[j].dateidx:
                        t[j].urlValid = False
                    else:
                        t[i].urlValid = False


    # aggregate data across the tasks
    task_date = {}
    for didx in range(14):
        task_date[didx] = False

    tasks_completed = 0
    bad_task_link = False
    bad_artifact_link = False
    bad_date = False
    tidx_old = tidx
    for tidx in range(tidx):
        # sort through the tasks
        if t[tidx].taskLinkValid and ~t[tidx].taskLinkMissing and \
                ~t[tidx].dateBadRange and t[tidx].dateValid and \
                t[tidx].urlValid and ~t[tidx].urlMissing:
            # here the task is valid - aggregate the data
            task_date[t[tidx].dateidx] = True
            if t[tidx].complete:
                tasks_completed = tasks_completed + t[tidx].percentWork

    # count the total number of contribution days
    contribution_gap = 0
    largest_gap = 0
    contribution_days = 0
    for didx in range(period_len):
        if task_date[didx]:
            contribution_days = contribution_days + 1
            contribution_gap = 0
        else:
            contribution_gap = contribution_gap + 1
            largest_gap = max(contribution_gap, largest_gap)

    # determine grade
    if tasks_completed < 1.0:
        grade = "Insufficient"
    elif largest_gap > 4 or contribution_days < 5:
        grade = "Novice"
    elif largest_gap > 3 or contribution_days < 6:
        grade = "Competent"
    else:
        grade = "Proficient"

    err_log = "Tasks Completed = " + str(tasks_completed/100) + ". " + err_log
    err_log = "Contribution days = " + str(contribution_days) + ". " + err_log
    if largest_gap>3:
        err_log = "Largest Gap = " + str(largest_gap) + ". " + err_log

    # store the results
    owsData['A' + str(output_line)] = name
    owsData['B' + str(output_line)] = ws[7][1].value
    owsData['C' + str(output_line)] = ws[8][1].value
    owsData['D' + str(output_line)] = ws[9][1].value
    owsData['E' + str(output_line)] = ws[10][1].value
    owsData['F' + str(output_line)] = grade
    owsData['G' + str(output_line)] = tasks_completed / 100
    owsData['H' + str(output_line)] = contribution_days
    owsData['I' + str(output_line)] = largest_gap
    owsData['J' + str(output_line)] = err_log
    output_line = output_line + 1
    iwb.close()
# vbs filename changed to .results.xlsx instead of results.xlsx

#ows = owb.create_sheet('badURLs')
#ws = iwb['badURLs']
#row = 0
#for url in badurl_list:
#    ws[row][0].value = url
#    ++row

owb.save("./.results.xlsx")


owb.close()

# output the bad url list
print()
print('Found ' + str(len(badurl_list)-1) + ' bad Task / Evidence URLs as given next')
print(badurl_list)